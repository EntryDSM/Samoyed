import requests
import json
from openpyxl import load_workbook
from flask import abort

from config import BASE_URL, STATUS_URL


def create_competition_status_excel(now_date, token):
    try:
        xlsx = load_workbook(f"static/competition_status.xlsx")
        sheet = xlsx.get_active_sheet()

        result = requests.get(STATUS_URL, params={"Authorization": token})

        response = json.loads(result.text)

        print(result.text)

        sheet.cell(4, 6, response['daejeon']['meister_applicant']['applicant_count'])
        sheet.cell(5, 6, response['daejeon']['social_applicant']['applicant_count'])
        sheet.cell(6, 6, response['daejeon']['common_applicant']['applicant_count'])

        sheet.cell(4, 7, response['nationwide']['meister_applicant']['applicant_count'])
        sheet.cell(5, 7, response['nationwide']['social_applicant']['applicant_count'])
        sheet.cell(6, 7, response['nationwide']['common_applicant']['applicant_count'])

        sheet.cell(7, 6, response['daejeon']['total_applicant_count'])
        sheet.cell(7, 8, response['nationwide']['total_applicant_count'])
        sheet.cell(8, 6, response['total_applicant_count'])

        sheet.cell(4, 8, response['daejeon']['meister_applicant']['competition_rate'])
        sheet.cell(5, 8, response['daejeon']['social_applicant']['competition_rate'])
        sheet.cell(6, 8, response['daejeon']['common_applicant']['competition_rate'])

        sheet.cell(4, 9, response['nationwide']['meister_applicant']['competition_rate'])
        sheet.cell(5, 9, response['nationwide']['social_applicant']['competition_rate'])
        sheet.cell(6, 9, response['nationwide']['common_applicant']['competition_rate'])

        sheet.cell(7, 8, response['daejeon']['total_competition_rate'])
        sheet.cell(7, 9, response['nationwide']['total_competition_rate'])
        sheet.cell(8, 8, response['total_competition_rate'])

        sheet.cell(12, 4, response['nationwide']['common_applicant']['141-150'])
        sheet.cell(13, 4, response['nationwide']['common_applicant']['131-140'])
        sheet.cell(14, 4, response['nationwide']['common_applicant']['121-130'])
        sheet.cell(15, 4, response['nationwide']['common_applicant']['111-120'])
        sheet.cell(16, 4, response['nationwide']['common_applicant']['101-110'])
        sheet.cell(17, 4, response['nationwide']['common_applicant']['91-100'])
        sheet.cell(18, 4, response['nationwide']['common_applicant']['81-90'])
        sheet.cell(19, 4, response['nationwide']['common_applicant']['71-80'])
        sheet.cell(20, 4, response['nationwide']['common_applicant']['-70'])

        sheet.cell(12, 5, response['nationwide']['meister_applicant']['141-150'])
        sheet.cell(13, 5, response['nationwide']['meister_applicant']['131-140'])
        sheet.cell(14, 5, response['nationwide']['meister_applicant']['121-130'])
        sheet.cell(15, 5, response['nationwide']['meister_applicant']['111-120'])
        sheet.cell(16, 5, response['nationwide']['meister_applicant']['101-110'])
        sheet.cell(17, 5, response['nationwide']['meister_applicant']['91-100'])
        sheet.cell(18, 5, response['nationwide']['meister_applicant']['81-90'])
        sheet.cell(19, 5, response['nationwide']['meister_applicant']['71-80'])
        sheet.cell(20, 5, response['nationwide']['meister_applicant']['-70'])

        sheet.cell(12, 6, response['nationwide']['social_applicant']['141-150'])
        sheet.cell(13, 6, response['nationwide']['social_applicant']['131-140'])
        sheet.cell(14, 6, response['nationwide']['social_applicant']['121-130'])
        sheet.cell(15, 6, response['nationwide']['social_applicant']['111-120'])
        sheet.cell(16, 6, response['nationwide']['social_applicant']['101-110'])
        sheet.cell(17, 6, response['nationwide']['social_applicant']['91-100'])
        sheet.cell(18, 6, response['nationwide']['social_applicant']['81-90'])
        sheet.cell(19, 6, response['nationwide']['social_applicant']['71-80'])
        sheet.cell(20, 6, response['nationwide']['social_applicant']['-70'])

        sheet.cell(12, 7, response['daejeon']['common_applicant']['141-150'])
        sheet.cell(13, 7, response['daejeon']['common_applicant']['131-140'])
        sheet.cell(14, 7, response['daejeon']['common_applicant']['121-130'])
        sheet.cell(15, 7, response['daejeon']['common_applicant']['111-120'])
        sheet.cell(16, 7, response['daejeon']['common_applicant']['101-110'])
        sheet.cell(17, 7, response['daejeon']['common_applicant']['91-100'])
        sheet.cell(18, 7, response['daejeon']['common_applicant']['81-90'])
        sheet.cell(19, 7, response['daejeon']['common_applicant']['71-80'])
        sheet.cell(20, 7, response['daejeon']['common_applicant']['-70'])

        sheet.cell(12, 8, response['daejeon']['meister_applicant']['141-150'])
        sheet.cell(13, 8, response['daejeon']['meister_applicant']['131-140'])
        sheet.cell(14, 8, response['daejeon']['meister_applicant']['121-130'])
        sheet.cell(15, 8, response['daejeon']['meister_applicant']['111-120'])
        sheet.cell(16, 8, response['daejeon']['meister_applicant']['101-110'])
        sheet.cell(17, 8, response['daejeon']['meister_applicant']['91-100'])
        sheet.cell(18, 8, response['daejeon']['meister_applicant']['81-90'])
        sheet.cell(19, 8, response['daejeon']['meister_applicant']['71-80'])
        sheet.cell(20, 8, response['daejeon']['meister_applicant']['-70'])

        sheet.cell(12, 9, response['daejeon']['social_applicant']['141-150'])
        sheet.cell(13, 9, response['daejeon']['social_applicant']['131-140'])
        sheet.cell(14, 9, response['daejeon']['social_applicant']['121-130'])
        sheet.cell(15, 9, response['daejeon']['social_applicant']['111-120'])
        sheet.cell(16, 9, response['daejeon']['social_applicant']['101-110'])
        sheet.cell(17, 9, response['daejeon']['social_applicant']['91-100'])
        sheet.cell(18, 9, response['daejeon']['social_applicant']['81-90'])
        sheet.cell(19, 9, response['daejeon']['social_applicant']['71-80'])
        sheet.cell(20, 9, response['daejeon']['social_applicant']['-70'])

        xlsx.save(f"static/competition_status_{now_date}.xlsx")

        return f"{BASE_URL}static/competition_status_{now_date}.xlsx"

    except Exception as e:
        return abort(500, e)
