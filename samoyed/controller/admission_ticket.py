import openpyxl
from openpyxl.drawing.image import Image
from flask import abort

from samoyed.config import BASE_URL
from samoyed.model.user import User
from samoyed.model.status import Status
from samoyed.model import session
from samoyed.controller import print_origin_school, print_is_daejeon, print_apply_type, create_str_cell_location


def create_admission_ticket(now_date):
    try:
        xlsx = openpyxl.load_workbook(f"samoyed/static/admission_ticket.xlsx")
        sheet = xlsx.get_active_sheet()

        users = session.query(User, Status).join(Status).all()

        for idx, user in enumerate(users):
            if idx % 4 == 0:
                img = openpyxl.drawing.image.Image(user.User.user_photo)

                unit_width = 58
                unit_inch_width = round((unit_width / 2.54) * 10, 0)

                unit_height = 64
                unit_inch_height = round((unit_height / 2.54) * 10, 0)

                img.width = unit_inch_width
                img.height = unit_inch_height

                sheet.add_image(img, create_str_cell_location(3 + (idx // 4 * 41), 1))
                sheet.cell(3 + (idx // 4 * 41), 5, user.Status.exam_code)
                sheet.cell(5 + (idx // 4 * 41), 5, user.User.name)
                sheet.cell(7 + (idx // 4 * 41), 5, print_origin_school(user.User.receipt_code, user.User.grade_type))
                sheet.cell(9 + (idx // 4 * 41), 5, print_is_daejeon(user.User.is_daejeon))
                sheet.cell(11 + (idx // 4 * 41), 5, print_apply_type(user.User.apply_type))
                sheet.cell(13 + (idx // 4 * 41), 5, user.User.receipt_code)

            elif idx % 4 == 1:
                img = openpyxl.drawing.image.Image(user.User.user_photo)

                unit_width = 58
                unit_inch_width = round((unit_width / 2.54) * 10, 0)

                unit_height = 64
                unit_inch_height = round((unit_height / 2.54) * 10, 0)

                img.width = unit_inch_width
                img.height = unit_inch_height

                sheet.add_image(img, create_str_cell_location(3 + (idx // 4 * 41), 8))
                sheet.cell(3 + (idx // 4 * 41), 12, user.Status.exam_code)
                sheet.cell(5 + (idx // 4 * 41), 12, user.User.name)
                sheet.cell(7 + (idx // 4 * 41), 12, print_origin_school(user.User.receipt_code, user.User.grade_type))
                sheet.cell(9 + (idx // 4 * 41), 12, print_is_daejeon(user.User.is_daejeon))
                sheet.cell(11 + (idx // 4 * 41), 12, print_apply_type(user.User.apply_type))
                sheet.cell(13 + (idx // 4 * 41), 12, user.User.receipt_code)

            elif idx % 4 == 2:
                img = openpyxl.drawing.image.Image(user.User.user_photo)

                unit_width = 58
                unit_inch_width = round((unit_width / 2.54) * 10, 0)

                unit_height = 64
                unit_inch_height = round((unit_height / 2.54) * 10, 0)

                img.width = unit_inch_width
                img.height = unit_inch_height

                sheet.add_image(img, create_str_cell_location(21 + (idx // 4 * 41), 1))
                sheet.cell(21 + (idx // 4 * 41), 5, user.Status.exam_code)
                sheet.cell(23 + (idx // 4 * 41), 5, user.User.name)
                sheet.cell(25 + (idx // 4 * 41), 5, print_origin_school(user.User.receipt_code, user.User.grade_type))
                sheet.cell(27 + (idx // 4 * 41), 5, print_is_daejeon(user.User.is_daejeon))
                sheet.cell(29 + (idx // 4 * 41), 5, print_apply_type(user.User.apply_type))
                sheet.cell(31 + (idx // 4 * 41), 5, user.User.receipt_code)

            elif idx % 4 == 3:
                img = openpyxl.drawing.image.Image(user.User.user_photo)

                unit_width = 58
                unit_inch_width = round((unit_width / 2.54) * 10, 0)

                unit_height = 64
                unit_inch_height = round((unit_height / 2.54) * 10, 0)

                img.width = unit_inch_width
                img.height = unit_inch_height

                sheet.add_image(img, create_str_cell_location(21 + (idx // 4 * 41), 8))
                sheet.cell(21 + (idx // 4 * 41), 12, user.Status.exam_code)
                sheet.cell(23 + (idx // 4 * 41), 12, user.User.name)
                sheet.cell(25 + (idx // 4 * 41), 12, print_origin_school(user.User.receipt_code, user.User.grade_type))
                sheet.cell(27 + (idx // 4 * 41), 12, print_is_daejeon(user.User.is_daejeon))
                sheet.cell(29 + (idx // 4 * 41), 12, print_apply_type(user.User.apply_type))
                sheet.cell(31 + (idx // 4 * 41), 12, user.User.receipt_code)

        xlsx.save(f"samoyed/static/admission_ticket_{now_date}.xlsx")

        return f"{BASE_URL}samoyed/static/admission_ticket_{now_date}.xlsx"

    except Exception as e:
        return abort(500, e)
