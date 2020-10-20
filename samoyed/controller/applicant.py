from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from flask import abort, send_file
from io import BytesIO

from samoyed.controller import (print_apply_type, print_additional_type, print_grade_type, print_is_daejeon,
                                print_sex, print_graduated_year, print_origin_school, print_student_number)
                                print_sex, print_graduated_year, print_origin_school, print_student_number,
                                print_social_type)
                                print_social_type, print_excel_num)
from samoyed.model import session
from samoyed.model.user import User
from samoyed.model.status import Status
from samoyed.model.ungraduated_application import UnGraduatedApplication
from samoyed.model.graduated_application import GraduatedApplication
from samoyed.model.calculated_score import CalculatedScore


def create_applicant_excel(date_time):
    try:
        xlsx = load_workbook(f"samoyed/static/applicant_info.xlsx")
        sheet = xlsx.get_active_sheet()

        users = session.query(User, CalculatedScore, Status).join(CalculatedScore).join(Status).filter(
            Status.is_final_submit == 1).all()

        for idx, user in enumerate(users):
            idx = idx + 2
            sheet.cell(idx, 1, user.Status.exam_code)
            sheet.cell(idx, 2, user.User.receipt_code)
            sheet.cell(idx, 2, print_excel_num(user.User.apply_type, user.User.is_daejeon, user.User.receipt_code))
            sheet.cell(idx, 3, print_apply_type(user.User.apply_type))
            sheet.cell(idx, 4, print_is_daejeon(user.User.is_daejeon))
            sheet.cell(idx, 5, print_additional_type(user.User.additional_type))
            sheet.cell(idx, 5, print_social_type(user.User.apply_type) + print_additional_type(user.User.additional_type))
            sheet.cell(idx, 6, user.User.name)
            sheet.cell(idx, 7, user.User.birth_date)
            sheet.cell(idx, 8, f"{user.User.address} {user.User.detail_address}")
            sheet.cell(idx, 9, user.User.applicant_tel)
            sheet.cell(idx, 10, print_sex(user.User.sex))
            sheet.cell(idx, 11, print_grade_type(user.User.grade_type))
            sheet.cell(idx, 12, print_graduated_year(user.User.receipt_code, user.User.grade_type))
            sheet.cell(idx, 13, print_origin_school(user.User.receipt_code, user.User.grade_type))
            sheet.cell(idx, 14, print_student_number(user.User.receipt_code, user.User.grade_type))
            sheet.cell(idx, 15, user.User.parent_name)
            sheet.cell(idx, 16, user.User.parent_tel)

            if user.User.grade_type == "UNGRADUATED":
                ungraduated_application = session.query(UnGraduatedApplication) \
                    .filter(UnGraduatedApplication.user_receipt_code == user.User.receipt_code).first()
                korean = ungraduated_application.korean
                social = ungraduated_application.social
                history = ungraduated_application.history
                math = ungraduated_application.math
                science = ungraduated_application.science
                tech_and_home = ungraduated_application.tech_and_home
                english = ungraduated_application.english

                for i in range(6):
                    sheet.cell(idx, 17 + 7 * i, korean[i])
                    sheet.cell(idx, 18 + 7 * i, social[i])
                    sheet.cell(idx, 19 + 7 * i, history[i])
                    sheet.cell(idx, 20 + 7 * i, math[i])
                    sheet.cell(idx, 21 + 7 * i, science[i])
                    sheet.cell(idx, 22 + 7 * i, tech_and_home[i])
                    sheet.cell(idx, 23 + 7 * i, english[i])

                sheet.cell(idx, 59, user.CalculatedScore.first_grade_score)
                sheet.cell(idx, 60, user.CalculatedScore.second_grade_score)
                sheet.cell(idx, 61, user.CalculatedScore.third_grade_score)
                sheet.cell(idx, 62, user.CalculatedScore.conversion_score)
                sheet.cell(idx, 63, ungraduated_application.volunteer_time)
                sheet.cell(idx, 64, user.CalculatedScore.volunteer_score)
                sheet.cell(idx, 65, ungraduated_application.full_cut_count)
                sheet.cell(idx, 66, ungraduated_application.late_count)
                sheet.cell(idx, 67, ungraduated_application.early_leave_count)
                sheet.cell(idx, 68, ungraduated_application.period_cut_count)

            elif user.User.grade_type == "GRADUATED":
                graduated_application = session.query(GraduatedApplication) \
                    .filter(GraduatedApplication.user_receipt_code == user.User.receipt_code).first()
                korean = graduated_application.korean
                social = graduated_application.social
                history = graduated_application.history
                math = graduated_application.math
                science = graduated_application.science
                tech_and_home = graduated_application.tech_and_home
                english = graduated_application.english

                for i in range(6):
                    sheet.cell(idx, 17 + 7 * i, korean[i])
                    sheet.cell(idx, 18 + 7 * i, social[i])
                    sheet.cell(idx, 19 + 7 * i, history[i])
                    sheet.cell(idx, 20 + 7 * i, math[i])
                    sheet.cell(idx, 21 + 7 * i, science[i])
                    sheet.cell(idx, 22 + 7 * i, tech_and_home[i])
                    sheet.cell(idx, 23 + 7 * i, english[i])

                sheet.cell(idx, 59, user.CalculatedScore.first_grade_score)
                sheet.cell(idx, 60, user.CalculatedScore.second_grade_score)
                sheet.cell(idx, 61, user.CalculatedScore.third_grade_score)
                sheet.cell(idx, 62, user.CalculatedScore.conversion_score)
                sheet.cell(idx, 63, graduated_application.volunteer_time)
                sheet.cell(idx, 64, user.CalculatedScore.volunteer_score)
                sheet.cell(idx, 65, graduated_application.full_cut_count)
                sheet.cell(idx, 66, graduated_application.late_count)
                sheet.cell(idx, 67, graduated_application.early_leave_count)
                sheet.cell(idx, 68, graduated_application.period_cut_count)

            sheet.cell(idx, 69, user.CalculatedScore.attendance_score)
            sheet.cell(idx, 70, user.CalculatedScore.final_score)
            sheet.cell(idx, 71, user.User.self_introduction)
            sheet.cell(idx, 72, user.User.study_plan)

        file = BytesIO(save_virtual_workbook(xlsx))

        return send_file(file,
                         as_attachment=True,
                         attachment_filename=f"applicant_info_{date_time}.xlsx")

    except Exception as e:
        abort(500, e)

    finally:
        session.close()